import webbrowser
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
import openpyxl
from tqdm import tqdm
import time

#--------------- Start Excel -----------------#
wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename='Allianz main sample.xlsm')
sheets = wb.sheetnames
ws = wb[sheets[0]]
#--------------- End Excel -------------------#

pbar = tqdm(total=46)
#----------------- Read Excel Rows -----------------#

try:	
	for i in range(4,50):
		try:
			index = str(i)
			sdId = ws['A'+index].value

			brand = ws['P'+index].value
			model = ws['T'+index].value
			serie = ws['U'+index].value
			serieSplit = serie.split(" ")
		
			age = ws['Z'+index].value
			zipcode = ws['M'+index].value
			nationality = ws['L'+index].value
		
			driver = webdriver.Chrome()
		
			driver.get("https://tarif.allianz.ch/asu_cdn/apps/asu_momf-gui/#/de/")
			wait = WebDriverWait(driver,20)
			time.sleep(5)

			driver.find_element_by_xpath("/html/body/div[2]/div/div/div/div[1]/button").click()
			modelElement = driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[1]/api-call/input")
			modelElement.click()
			time.sleep(1)
			modelElement.send_keys(brand)
			time.sleep(1)
			modelElement.send_keys(" "+model)
			time.sleep(1)
			modelElement.send_keys(" "+serieSplit[0])
			time.sleep(1)
			modelElement.send_keys(" "+serieSplit[1])
			time.sleep(1)
			modelElement.send_keys(" "+serieSplit[2])
			time.sleep(1)
		
			modelElement.send_keys(Keys.ENTER)
			time.sleep(1)
			yearElement= driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[2]/api-call/drop-down-select/div/input[2]") 
			yearElement.click()
			time.sleep(1)
			yearElement.send_keys(Keys.ARROW_DOWN)
			driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[2]/api-call/drop-down-select/div/ul/li[1]/ul/li[2]").click()
		
		
			ageElement = driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[3]/api-call/input").send_keys(str(age))
			zipCodeElement = driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[3]/span/api-call/input")
			time.sleep(1)
			zipCodeElement.clear()
			time.sleep(1)
			zipCodeElement.send_keys(str(zipcode))
			time.sleep(1)
			zipCodeElement.send_keys(Keys.ARROW_DOWN)
			zipCodeElement.send_keys(Keys.RETURN)
			nat = driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[4]/api-call/input")
			nat.clear()
			time.sleep(1)
			nat.send_keys(nationality)
			time.sleep(1)
			nat.send_keys(Keys.ARROW_DOWN)
			nat.send_keys(Keys.RETURN)
			driver.find_element_by_xpath("/html/body/div[2]/div/div/form/div/div/p[5]/button").click()



			time.sleep(5)
			
			thirdValue = driver.find_element_by_xpath("/html/body/div[2]/div/div[1]/table/tbody[1]/tr[4]/td[2]/div/div[2]/div[1]/strong").text
			thirdValue = thirdValue.replace('CHF','')
			thirdValue = thirdValue.replace('.','')
			thirdValue = thirdValue.replace('*','')
			ws["Y"+index] = thirdValue.strip()
			driver.find_element_by_xpath("/html/body/div[2]/div/div[1]/table/tbody[3]/tr[1]/td[1]/checkbox").click()
			time.sleep(3)
			secondValue = driver.find_element_by_xpath("/html/body/div[2]/div/div[1]/table/tbody[1]/tr[4]/td[2]/div/div[2]/div[1]/strong").text
			secondValue = secondValue.replace('CHF','')
			secondValue = secondValue.replace('.','')
			secondValue = secondValue.replace('*','')
			ws["X"+index] = secondValue.strip()
			
			
			driver.find_element_by_xpath("/html/body/div[2]/div/div[1]/table/tbody[4]/tr[1]/td[1]/checkbox").click()
			time.sleep(3)
			firstValue = driver.find_element_by_xpath("/html/body/div[2]/div/div[1]/table/tbody[1]/tr[4]/td[2]/div/div[2]/div[1]/strong").text
			firstValue = firstValue.replace('CHF','')
			firstValue = firstValue.replace('.','')	
			firstValue = firstValue.replace('*','')	
			ws["W"+index] = firstValue.strip()
			
			time.sleep(1)
		
		
		except:
			print("Shit Happened")
			pass
			#driver.close()
			#driver.quit()

		finally:
			pbar.update(1)
			driver.close()
			driver.quit()

	pbar.close()
except:
	print "General Error"
finally:
	wb.save("Allianz main sample.xlsm")
	
